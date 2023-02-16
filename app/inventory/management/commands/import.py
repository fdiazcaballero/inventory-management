from django.core.management.base import BaseCommand
import openpyxl
from inventory.models import Location, Staff, Ingredient, Recipe, RecipeIngredient, Modifier, ModifierOption, Menu
from django.contrib.auth.models import User
from django.db import transaction


class Command(BaseCommand):
    help = 'Import data from file'

    def handle(self, *args, **options):
        wb_obj = openpyxl.load_workbook('inventory/management/commands/Weird_Salads_Data_Export.xlsx')
        sheet = wb_obj.active
        for sheet in wb_obj.worksheets:
            match sheet.title:
                case 'locations':
                    self.import_locations(sheet)
                case 'staff':
                    self.import_staff(sheet)
                case 'ingredients':
                    self.import_ingredients(sheet)
                case 'recipes':
                    self.import_recipes(sheet)
                case 'modifiers':
                    self.import_modifiers(sheet)
                case 'menus':
                    self.import_menus(sheet)

        self.stdout.write(self.style.SUCCESS('IMPORT SUCCESS!!'))

    def import_staff(self, sheet):
        with transaction.atomic():
            # This below code executes inside a transaction.
            for row in sheet.iter_rows(min_row=2, values_only=True):
                staff = Staff.objects.filter(staff_id=row[0]).first()
                if not staff:
                    user = User(username='_'.join(row[1].split()), is_active='False')
                    user.save()
                    staff = Staff(staff_id=row[0], name=row[1], dob=row[2], role=row[3], iban=row[4], bic=row[5], user=user)
                    staff.save()
                staff.location.add(Location.objects.filter(location_id=row[6]).first())
                staff.save()

    def import_locations(self, sheet):
        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):
                location = Location(name=row[1], address=row[2])
                location.save()

    def import_ingredients(self, sheet):
        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):
                ingredient = Ingredient(name=row[1], unit=row[2], cost=row[3])
                ingredient.save()

    def import_recipes(self, sheet):
        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):
                recipe, created = Recipe.objects.get_or_create(name=row[1])
                recipe_ingredient = RecipeIngredient(recipe=recipe, quantity=row[2], ingredient=Ingredient.objects.filter(ingredient_id=row[3]).first())
                recipe_ingredient.save()

    def import_modifiers(self, sheet):
        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):
                modifier, created = Modifier.objects.get_or_create(name=row[1])
                modifier_option = ModifierOption(option=row[2], price=row[3], modifier=modifier)
                modifier_option.save()

    def import_menus(self, sheet):
        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):
                menu = Menu(
                    recipe=Recipe.objects.filter(recipe_id=row[0]).first(),
                    location=Location.objects.filter(location_id=row[1]).first(),
                    price=row[2],
                    modifier=Modifier.objects.filter(modifier_id=row[3]).first() if row[3] else None
                )
                menu.save()
